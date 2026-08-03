"""Generate AI-ManJu project brief Excel from the template image."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "AI漫剧立项分析"

# Columns exactly matching the reference image
columns = [
    "立项",
    "锚定一个行业",
    "行业交流频繁的专业词汇",
    "核心业务线",
    "已查明真实需求点",
    "现有产品案例",
    "核心痛点",
    "信息来源备注",
    "确定项目名",
]

# One comprehensive row for "国内 AI 漫剧" industry
row = [
    # 立项
    ("以 AI 漫剧为切入点，打造\"剧本→分镜→生图→生视频→配音→剪辑→投流\"\n"
     "全流程工业化生产平台，先服务中小团队与一人公司(OPC)，\n"
     "再向 IP 方、平台方、B 端品牌定制延伸。\n"
     "本期目标：把单分钟制作成本压到 800 元以内、把单集周期压到 5 天以内、\n"
     "把跨工具工作流从 11 步压到 5 步以内。"),
    # 锚定一个行业
    "国内 AI 漫剧 / AI 微短剧内容生产行业\n"
    "(网文/漫画 IP × 生成式视频大模型 × 短剧分账生态)",
    # 行业交流频繁的专业词汇
    "AI 漫剧、表情包漫剧、AI 仿真人剧、微短剧\n"
    "分镜、镜一致性、人物建模一致性、声画同出、抽卡\n"
    "文生图、图生视频、视频续写、首尾帧控制\n"
    "网文 IP、番茄 IP、IP 矩阵、爆款率、破亿率\n"
    "投流、推流、万播分账、IAA、IAP、IAAP、会员分账\n"
    "流水线工作流、一站式平台、一人公司 OPC、工业化产能\n"
    "算力成本、人力成本、版权交易、漫剧带消费\n"
    "男频/女频、玄幻仙侠、末日求生、系统流、甜宠复仇",
    # 核心业务线
    "① AI 漫剧制作服务：从剧本到成片的端到端承制\n"
    "② 工作流/工具链：跨模型(可灵/即梦/通义万相/Vidu/Sora2)统一调度\n"
    "③ IP 改编与孵化：对接阅文/番茄/中文在线等网文 IP 库\n"
    "④ 平台分发与投流：抖音/快手/红果/B 站漫剧扶持计划对接\n"
    "⑤ B 端定制：品牌广告、出版机构、游戏公司定制 AI 漫剧\n"
    "⑥ 版权交易与 IP 衍生：漫剧带内容、漫剧带消费长期变现\n"
    "⑦ 算力套餐与计费 SaaS：按分钟/按集数计费、按角色锁定计费",
    # 已查明真实需求点
    "① 人物/场景跨镜一致性差，被业内称为\"最大技术痛点\"\n"
    "② 分镜抽卡随机性高，单段常需 20+ 次重试，浪费算力与时间\n"
    "③ 长叙事稳定性差，超过 1 分钟即崩坏，专业团队仍返工\n"
    "④ 跨工具工作流断裂：剧本/分镜/生图/视频/剪辑不在同一平台\n"
    "⑤ 微表情、口型同步、肢体动作僵硬，AI 仿真人剧\"一眼假\"\n"
    "⑥ 单分钟成本希望从 2000-5000 元压到 1000 元以下\n"
    "⑦ 制作周期希望从 3 个月压到 7-30 天\n"
    "⑧ 算力成本核算与计费模型粗糙，无统一账单与对账\n"
    "⑨ 投流 ROI 不稳定，盈利模式高度依赖平台分账\n"
    "⑩ 一人公司 / 5-15 人小团队缺协作流程与角色分工模板\n"
    "⑪ IP 版权归属模糊，与原 IP 方/作者分账规则不清\n"
    "⑫ 同质化严重，免费模式下\"万播分账\"有效曝光占比低",
    # 现有产品案例
    "工具/平台：即梦 AI(字节)、可灵 AI(快手)、通义万相 Wan2.5(阿里)、\n"
    "Vidu(生数)、Sora2(OpenAI)、Runway Gen4、纳米漫剧流水线(360)\n"
    "内容方：阅文漫剧助手、妙笔通鉴、版权助手、中文在线《明日周一》、\n"
    "灵境万维《我在末世开超市》(1200 万收入/15 万成本)、\n"
    "澄文影业(王娟)、锐和影视(罗时海)、甚妙《灵魂操纵师妙灵》、\n"
    "成都发光橙子、城市传奇(厦门)\n"
    "平台方：抖音漫剧扶持计划、星光/辰星计划、快手漫剧专项基金、\n"
    "B 站 AI 漫剧场\"觉醒计划\"(最高 80% 分成)、番茄 IP 改编合作\n"
    "数据/榜单：DataEye 短剧数据、灯塔专业版",
    # 核心痛点
    "技术层：人物一致性差、抽卡率高、长叙事崩坏、微表情僵硬、\n"
    "不支持真人人脸参考与 IP 形象锁定、Seedance 类模型排队与降智\n"
    "商业层：爆款率极低(破亿率 0.117%)、内容同质化、内卷压缩利润、\n"
    "下游分账不稳定、男频题材广告植入受限\n"
    "工作流层：跨工具链断裂、人力仍是大头(编剧+导演+生图生视频+配音)、\n"
    "算力与计费粗糙、无统一角色库/分镜库/资产库\n"
    "合规层：AI 生成内容标识、版权归属、网文 IP 二次改编授权链路不清晰\n"
    "供给层：供需失衡、头部平台议价权强、上游承制公司利润空间被挤压",
    # 信息来源备注
    "1) 中国网新闻 2026.06《AI 漫剧，视频内容市场\"新变量\"》\n"
    "2) 界面新闻《即梦和可灵，能不能接住 AI 短剧风口？》\n"
    "3) 澎湃新闻《AI 漫剧快速成长，各平台疯抢新赛道》\n"
    "4) 厦门日报 2025.12《AI 微短剧站上新风口 多家厦企\"跑步入场\"》\n"
    "5) 经济参考报 2026.01《低成本 AI 漫剧告别野蛮生长》\n"
    "6) 光明网/百度百家号 2025.12《AI 短剧爆发式增长 百亿级市场加速成势》\n"
    "7) 腾讯新闻《漫剧爆发背后：微短剧行业的一场降本试验》\n"
    "8) 企查查 2025 漫剧企业注册数据(8.02 万家/+37.1%)\n"
    "9) DataEye 短剧数据(2026.02 在播 12.78 万部)\n"
    "10) 东吴证券、艾媒咨询、深度科技研究院相关研报\n"
    "采集时间：2026-08-01；以上来源均可回溯，部分为公开报道而非一手数据",
    # 确定项目名（已定稿）
    "【已确定】漫镜工场 (ManJu Studio)\n"
    "定位：一站式 AI 漫剧工业化生产平台（平台+工作流+服务三层）\n"
    "产品名确认时间：2026-08-02（PRD v1.0 定稿）\n"
    "候选备选：灵境漫剧云 / 秒镜 AI / 橙镜漫剧 / 分镜师\n"
    "最新配套：PRD v1.3 + 流程方案 v1.1 + 技术栈说明书 v1.2",
]

# ---- Style ----
header_fill = PatternFill("solid", fgColor="1F3A5F")
header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
cell_font = Font(name="Microsoft YaHei", size=10, color="1A1A1A")
title_font = Font(name="Microsoft YaHei", size=14, bold=True, color="1F3A5F")
note_font = Font(name="Microsoft YaHei", size=9, italic=True, color="555555")
wrap_align = Alignment(wrap_text=True, vertical="top", horizontal="left")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(style="thin", color="B5B5B5")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Title row
ws.cell(row=1, column=1, value="国内 AI 漫剧行业 · 立项分析表").font = title_font
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
ws.row_dimensions[1].height = 28

# Subtitle / meta row
ws.cell(row=2, column=1, value=(
    "立项人：B 哥  |  日期：2026-08-03（项目名已定稿）  |  数据基准：2025 全年 + 2026 上半年  |  "
    "采集方式：公开报道、券商研报、平台公开数据；非一手数据均标来源"
)).font = note_font
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
ws.row_dimensions[2].height = 18

# Header row
for i, name in enumerate(columns, start=1):
    c = ws.cell(row=3, column=i, value=name)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_align
    c.border = border
ws.row_dimensions[3].height = 32

# Data row
for i, value in enumerate(row, start=1):
    c = ws.cell(row=4, column=i, value=value)
    c.font = cell_font
    c.alignment = wrap_align
    c.border = border
ws.row_dimensions[4].height = 380

# Column widths (proportional to content density)
widths = [22, 22, 26, 26, 32, 32, 30, 28, 26]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Footer
ws.cell(row=6, column=1, value=(
    "使用说明：本表依据公开行业素材整理，所有外部数据已标注来源名称与采集时间；\n"
    "凡未在\"信息来源备注\"中列出的具体百分比/播放量/收入数据，均视为待核验假设，\n"
    "不可直接作为客户稿或对外引用的\"一手数据\"。下一步建议：先挑 1 个候选项目名做 3 天试点。"
)).font = note_font
ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=len(columns))
ws.row_dimensions[6].height = 36
ws.cell(row=6, column=1).alignment = wrap_align

# Freeze header
ws.freeze_panes = "A4"

out = r"C:\Users\李俊麟\Desktop\workbench\AI漫剧项目\AI漫剧行业立项分析表.xlsx"
wb.save(out)
print("SAVED:", out)